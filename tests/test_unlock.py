import sys
from pathlib import Path

# Ajouter src/ au chemin d'importation
src_path = Path(__file__).resolve().parent.parent / "src"
if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

import openpyxl
from unlock import remove_excel_protection, inspect_excel_file


def test_sheet_protection_removal(tmp_path: Path):
    test_file = tmp_path / "protected_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Valeur Confidentielle"
    ws.protection.sheet = True
    ws.protection.password = "password123"
    wb.save(test_file)

    # Déverrouillage
    success = remove_excel_protection(test_file)
    assert success is True
    assert test_file.exists()

    # Vérification
    wb_unlocked = openpyxl.load_workbook(test_file)
    ws_unlocked = wb_unlocked["Data"]
    assert ws_unlocked["A1"].value == "Valeur Confidentielle"
    assert ws_unlocked.protection.sheet is False


def test_workbook_structure_protection_removal(tmp_path: Path):
    test_file = tmp_path / "protected_wb.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Test Structure"
    wb.security.lockStructure = True
    wb.security.lockWindows = True
    wb.save(test_file)

    # Déverrouillage
    success = remove_excel_protection(test_file)
    assert success is True

    # Vérification
    wb_unlocked = openpyxl.load_workbook(test_file)
    assert getattr(wb_unlocked.security, "lockStructure", None) is not True
    assert getattr(wb_unlocked.security, "lockWindows", None) is not True


def test_selective_sheet_unlocking(tmp_path: Path):
    test_file = tmp_path / "multi_sheets.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Feuille_1"
    ws1.protection.sheet = True
    ws1.protection.password = "p1"

    ws2 = wb.create_sheet("Feuille_2")
    ws2.protection.sheet = True
    ws2.protection.password = "p2"

    wb.security.lockStructure = True
    wb.save(test_file)

    # Inspection
    elements = inspect_excel_file(test_file)
    target_f1 = [e["target"] for e in elements if "Feuille_1" in e["title"]][0]

    # Déverrouiller uniquement Feuille 1
    success = remove_excel_protection(test_file, targets_to_unlock=[target_f1])
    assert success is True

    # Vérification que Feuille 1 est déverrouillée et Feuille 2 reste verrouillée
    wb_check = openpyxl.load_workbook(test_file)
    assert wb_check["Feuille_1"].protection.sheet is False
    assert wb_check["Feuille_2"].protection.sheet is True
    assert getattr(wb_check.security, "lockStructure", None) is True


def test_backup_option(tmp_path: Path):
    test_file = tmp_path / "with_backup.xlsx"
    wb = openpyxl.Workbook()
    wb.save(test_file)

    success = remove_excel_protection(test_file, make_backup=True)
    assert success is True

    backup_file = tmp_path / "with_backup.xlsx.bak"
    assert backup_file.exists()
